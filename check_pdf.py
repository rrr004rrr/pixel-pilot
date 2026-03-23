"""
check_pdf.py — 比對任務清單與已下載 PDF

用法：
    python check_pdf.py <xlsx路徑>
    python check_pdf.py <xlsx路徑> <PDF資料夾>
    python check_pdf.py "D:/TODOPDF/任务列表.xlsx"
    python check_pdf.py "D:/TODOPDF/任务列表.xlsx" "D:/TODOPDF"

輸出：
    - 終端機顯示統計摘要
    - 同目錄產生 check_result_<時間戳>.xlsx（已下載 / 未下載 兩個工作表）

依賴：
    pip install openpyxl pdfplumber
"""

import sys
import re
import warnings
from datetime import datetime
from pathlib import Path

DEFAULT_PDF_FOLDER = r"D:\TODOPDF"
_INVALID = re.compile(r'[\\/:*?"<>|]')


def normalize(name: str) -> str:
    """統一化名稱：去除首尾空白、轉小寫，方便比對。"""
    return name.strip().lower()


def extract_title(pdf_path: Path) -> str | None:
    """讀取 PDF 第一行文字作為標題，失敗回傳 None。"""
    try:
        import pdfplumber
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            with pdfplumber.open(str(pdf_path)) as pdf:
                text = pdf.pages[0].extract_text() or ""
        first_line = text.strip().splitlines()[0].strip() if text.strip() else ""
        return first_line or None
    except Exception:
        return None


def build_pdf_title_set(pdf_folder: Path) -> set[str]:
    """掃描資料夾內所有 PDF，回傳標題集合（normalized）。"""
    pdfs = list(pdf_folder.glob("*.pdf"))
    total = len(pdfs)
    titles = set()
    for i, p in enumerate(pdfs, 1):
        if i % 200 == 0 or i == total:
            print(f"  讀取 PDF 進度：{i}/{total}...", end="\r")
        title = extract_title(p)
        if title:
            titles.add(normalize(title))
    print()
    return titles


def load_tasks(xlsx_path: Path) -> list[dict]:
    """讀取 xlsx，回傳每列的欄位字典。"""
    import openpyxl
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:]]


def check(xlsx_path: Path, pdf_folder: Path):
    print(f"PDF 資料夾：{pdf_folder}")
    print(f"任務清單：{xlsx_path.name}")
    print()

    print("掃描 PDF 檔案內容（讀取標題）...")
    pdf_titles = build_pdf_title_set(pdf_folder)
    print(f"  成功讀取 {len(pdf_titles)} 個 PDF 標題")
    print()

    tasks = load_tasks(xlsx_path)
    if not tasks:
        print("無法讀取任務清單")
        return
    print(f"任務清單筆數：{len(tasks)}")

    downloaded = []
    not_downloaded = []

    for task in tasks:
        name   = str(task.get("任务名称") or "").strip()
        number = str(task.get("任务编号") or "").strip()
        if not name:
            continue
        # PDF 第一行格式為「任务编号 任务名称」
        key = normalize(f"{number} {name}" if number else name)
        if key in pdf_titles:
            downloaded.append(task)
        else:
            not_downloaded.append(task)

    print(f"  已下載：{len(downloaded)}")
    print(f"  未下載：{len(not_downloaded)}")
    print()

    # 輸出結果 xlsx
    import openpyxl
    from openpyxl.styles import PatternFill, Font

    wb_out = openpyxl.Workbook()
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    RED   = PatternFill("solid", fgColor="FFC7CE")
    BOLD  = Font(bold=True)
    headers = list(tasks[0].keys()) if tasks else []

    def write_sheet(ws, rows, fill):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = BOLD
        for task in rows:
            ws.append([task.get(h) for h in headers])
            for cell in ws[ws.max_row]:
                cell.fill = fill

    ws1 = wb_out.active
    ws1.title = f"已下載({len(downloaded)})"
    write_sheet(ws1, downloaded, GREEN)

    ws2 = wb_out.create_sheet(f"未下載({len(not_downloaded)})")
    write_sheet(ws2, not_downloaded, RED)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = xlsx_path.parent / f"check_result_{ts}.xlsx"
    wb_out.save(str(out_path))
    print(f"結果已儲存：{out_path.name}")


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    args = sys.argv[1:]
    if not args:
        print("用法：python check_pdf.py <xlsx路徑> [PDF資料夾]")
        sys.exit(1)

    xlsx_path = Path(args[0]).expanduser().resolve()
    if not xlsx_path.exists():
        print(f"找不到檔案：{xlsx_path}")
        sys.exit(1)

    pdf_folder = Path(args[1]).expanduser().resolve() if len(args) > 1 else Path(DEFAULT_PDF_FOLDER)
    if not pdf_folder.exists():
        print(f"找不到資料夾：{pdf_folder}")
        sys.exit(1)

    check(xlsx_path, pdf_folder)
